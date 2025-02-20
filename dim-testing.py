import os
import sys
import json
import time
import joblib
import logging
import pandas as pd
import tkinter as tk 
from itertools import permutations
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tkinter import filedialog, ttk, font

# Get the user's Downloads folder path
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")

# File to store the tolerances
tol_file = 'tolerances.json'

def load_values():
    """Load the last used values from a JSON file."""
    if os.path.exists(tol_file):
        with open(tol_file, 'r') as f:
            return json.load(f)
    return {"length": 0.2, "width": 0.2, "height": 0.2}  # Default values

def store_values():
    """Store the current entry values to a JSON file."""
    tolerances = {
        "length": float(length_tol_entry.get()),
        "width":  float(width_tol_entry.get()),
        "height": float(height_tol_entry.get())
    }
    
    with open(tol_file, 'w') as f:
        json.dump(tolerances, f)

def validate_numeric_input(action, value_if_allowed):
    # Validate numeric input (allows decimal values)
    if action != '1':  # If action is not '1', it means it's not an insert action
        return True
    try:
        float(value_if_allowed)
        return True
    except ValueError:
        return False

def save_excel_file(output_df, excel_file, stdout):
    # Reset stdout back to original
    sys.stdout = stdout  

    # Extract the file name without the folder name
    excel_filename = os.path.basename(excel_file)

    while True:
        try:
            # Try to save the DataFrame to Excel
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='Results')

            # Load the workbook to apply formatting
            wb = load_workbook(excel_file)
            ws = wb.active  # Get active sheet

            # Define colors for highlighting
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            red_font = Font(color="9C0006", bold=False)  # Dark red (unbolded)

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            green_font = Font(color="006100", bold=False)  # Dark green (unbolded)

            # Get tolerance values
            tolerances = load_values()

            # Find column indices for ΔLength, ΔWidth, and ΔHeight
            col_indices = {col: idx + 1 for idx, col in enumerate(output_df.columns) if col in ['ΔLength', 'ΔWidth', 'ΔHeight']}

            # Apply conditional formatting
            for row_idx, row in enumerate(output_df.itertuples(), start=2):  # Start from row 2 (skip header)
                for col, col_idx in col_indices.items():
                    value = getattr(row, col)
                    tolerance = tolerances[col.replace('Δ', '').lower()]  # Get tolerance for Length, Width, or Height

                    # Apply red if outside tolerance, green if within tolerance
                    if abs(value) > tolerance:
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill
                        ws.cell(row=row_idx, column=col_idx).font = red_font
                    else:
                        ws.cell(row=row_idx, column=col_idx).fill = green_fill
                        ws.cell(row=row_idx, column=col_idx).font = green_font

            # Save the formatted Excel file
            wb.save(excel_file)

            break  # Exit the loop once the file is saved
        except PermissionError:
            print(f"\nUnable to save results to an Excel file because \"{excel_filename}\" is currently open. Please close the file to proceed.")
            user_input = input("Are you ready to retry? (Y/n): ").strip().lower()
            if user_input == 'n':
                print("Exiting without saving.")
                break
            elif user_input == 'y':
                print("Retrying to save the file...")
                time.sleep(1)  # Wait for a second before retrying
            else:
                print("Invalid input. Please type 'Y' or 'n'.")

def load_file():
    file_path = filedialog.askopenfilename(filetypes=[("Log Files", "*.log"), ("All Files", "*.*")])
    if file_path:
        # Extract the last few parts of the path
        truncated_path = os.path.join(*file_path.split(os.sep)[-6:])  # Adjust the number as needed to control how much of the path to show
        log_file_entry.delete(0, "end")  # Deletes entry if present
        log_file_entry.insert(0, truncated_path)  # Inserts the entry with the truncated path
        return file_path
    else:
        log_file_entry.config(text="No file selected.")
        return None
    
def calculate_min_difference(row):
    # Specify Actual vs Result dimensions
    actual_dims = (row['Length Actual'], row['Width Actual'], row['Height Actual'])
    result_dims = (row['Length'], row['Width'], row['Height'])
    
    # Generate all permutations (rotations) of the expected dimensions
    rotations = list(permutations(actual_dims))
    
    # Initialize parameters for finding the best rotation
    min_difference = None
    best_rotation = None
    
    for rotation in rotations:
        # Calculate the difference for this rotation
        difference = [result_dims[i] - rotation[i] for i in range(3)]
        
        # Compute the total absolute difference
        total_difference = sum(abs(diff) for diff in difference)
        
        # Update minimum difference and best rotation
        if min_difference is None or total_difference < min_difference:
            min_difference = total_difference
            best_rotation = rotation
    
    # Format differences to .2f
    formatted_difference = [f"{diff:.2f}" for diff in [result_dims[i] - best_rotation[i] for i in range(3)]]
    
    return pd.Series(formatted_difference, index=['Difference Length', 'Difference Width', 'Difference Height'])

def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)

def save_selected_boxes(selected_boxes, file_path=resource_path("selected_boxes.json")):
    """Save selected boxes to a file."""
    with open(file_path, 'w') as f:
        json.dump(selected_boxes, f)

def load_selected_boxes(file_path=resource_path("selected_boxes.json")):
    """Load selected boxes from a file."""
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            return json.load(f)
    return None

def filter_boxes(df, checkboxes):
    global selected_boxes, filtered_boxes  # Declare filtered_df as a global variable
    # Get the selected box sizes
    selected_boxes = [box for box, var in checkboxes.items() if var.get() == '1']

        # Save the selected boxes to a file
    save_selected_boxes(selected_boxes)
    
    # Filter the DataFrame to include only the selected box sizes
    filtered_boxes = df[df['Box'].isin(selected_boxes)]

def setup_logging(folder):
    # Only set up logging if an error is detected
    log_file = os.path.join(folder, "error.log")
    logging.basicConfig(filename=log_file, level=logging.ERROR)

    return logging, log_file

def save_selected_boxes(selected_boxes, file_path=resource_path("selected_boxes.json")):
    """Save selected boxes to a file."""
    with open(file_path, 'w') as f:
        json.dump(selected_boxes, f)

def parse_log():
    try:
        filter_boxes(box_df, checkboxes)

        original_stdout = sys.stdout  # Save the original stdout
        print("Running script...")
        # Define the output directory and filename
        if not log_file_entry.get():
            print("No valid file entered....")
            return

        # Get the base name (filename with extension)
        base_name = os.path.basename(log_file_entry.get())

        # Split the base name into filename and extension
        file_name, _ = os.path.splitext(base_name)

        # Create a directory with filename
        output_path = os.path.join(downloads_folder, f"output/{file_name}")
        os.makedirs(output_path, exist_ok=True)

        summary_file = os.path.join(output_path, "summary.txt")

        # Redirect stdout to the summary file
        with open(summary_file, 'w') as file:
            sys.stdout = file  # Redirect stdout to the file

            # To load the model
            knn = joblib.load(resource_path('model.joblib'))

            # Create a dataframe of all the measurements from the log file
            meas_df = pd.read_csv(log_file_entry.get(), sep=';')  # new box dimensions
  
            # If "Status 3" exists
            if "Status 3" in meas_df.columns:
                meas_df = meas_df[meas_df["Status 3"] != 0]

            # If "DIM State 3" exists
            elif "DIM State 3" in meas_df.columns:
                meas_df = meas_df[meas_df["DIM State 3"] != 0]
                
            else:
                # Neither column exists
                print("Neither 'Status 3' nor 'DIM State 3' columns are present in the DataFrame.")

            meas_df[['Length', 'Width', 'Height']] = (meas_df[['Length', 'Width', 'Height']]).round(1) # rounds all L, W, H to nearest tenth (5.799999 -> 5.8)

            # Predicts the actual dimensions based on measured data
            meas_df.loc[:,'Box'] = knn.predict((meas_df[['Length', 'Width', 'Height']]))
            meas_df = meas_df[meas_df['Box'].isin(selected_boxes)] # drops unselected boxes

            # Merge the DataFrames based on the label column, preserving the order of meas_df
            merged_df = meas_df.merge(filtered_boxes, on='Box', suffixes=(' Measured', ' Actual'), how='left')

            # Rename the columns for cleaner look in Excel
            merged_df = merged_df.rename(columns={
                'Length Measured': 'Length',
                'Width Measured' : 'Width',
                'Height Measured': 'Height'
            })

            # Subtract the corresponding features
            for col in ['Length', 'Width', 'Height']:
                merged_df[f'Δ{col}'] = merged_df[f'{col}'] - merged_df[f'{col} Actual']

            # Align Actual vs Result dimensions (i.e. "5.2x6.2x2.0" would be "6x5x2" box but calculated difference would be "5x6x2")
            merged_df[[f'Δ{col}' for col in ['Length', 'Width', 'Height']]] = merged_df.apply(calculate_min_difference, axis=1).astype(float)

            # Select only the columns containing the differences
            difference_df = merged_df[[f'Δ{col}' for col in ['Length', 'Width', 'Height']]]

            # Calculate the frequency each column is out of spec (greater than ± 0.2)
            count_ole = sum((round(difference_df['ΔLength'].abs(), 1) > tolerances["length"]))
            count_owi = sum((round(difference_df['ΔWidth'].abs(), 1) >  tolerances["width"]) )
            count_ohi = sum((round(difference_df['ΔHeight'].abs(), 1) > tolerances["height"]))
                
            # Calculate the number of rows with populated dimensions
            total_rows = difference_df.shape[0] 

            # Check whether any dimensions are out of spec
            if count_ole > 0 or count_owi > 0 or count_ohi > 0:

                # Print the occurrences each time length, width, and height is off 
                print(f"Length is off:  {count_ole} out of {total_rows} time(s)")
                print(f"Width  is off:  {count_owi} ouf of {total_rows} time(s)")
                print(f"Height is off:  {count_ohi} out of {total_rows} time(s)\n")

            else:
                print(f"All populated dimensions are within spec!") 

            # Filter the dimensions that failed
            mask =  (round(merged_df['ΔLength'].abs(), 1) > tolerances["length"]) | \
                    (round(merged_df['ΔWidth'].abs(),  1) > tolerances["width"] ) | \
                    (round(merged_df['ΔHeight'].abs(), 1) > tolerances["height"])

            # Filter out the bad dimensions and count number of occurrences
            filtered_df = merged_df[mask]
            total_bad = filtered_df.shape[0]

            # Count occurrences of each unique set of box failures
            failure_counts = filtered_df.groupby(['Box']).size()

            # Sort the filtered DataFrame by the 'Index' column
            sorted_failed_boxes = filtered_df[['Index', 'Length', 'Width', 'Height', 'Box']].sort_values(by=['Box', 'Index'])

            # Convert the sorted DataFrame to a string without the default index
            failed_boxes = sorted_failed_boxes.to_string(index=False)

            # Print boxes that fail
            for label, count in failure_counts.items():
                print(f"Box {label} is out of spec {count} time(s)")
                
            # Set the display option to expand the column width
            pd.set_option('display.max_colwidth', None)

            # Calculate and print the success rate; print failed boxes if applicable
            success_rate = (total_rows - total_bad) / total_rows * 100
            print(f"\n{total_bad} out of {total_rows} boxes failed: {success_rate:.2f}% success rate\n", f"\nFailed boxes:\n{failed_boxes}" if total_bad else "")

            # Create an output DataFrame to export to an Excel file with only valid dimensions
            output_df = merged_df[['Index', 'Length', 'Width', 'Height', 'Box', 'ΔLength', 'ΔWidth', 'ΔHeight', 'DIM State 1', 'DIM State 2', 'DIM State 3']]

            output_df = output_df.rename(columns={ 
                'DIM State 1': 'State 1',
                'DIM State 2': 'State 2',
                'DIM State 3': 'State 3'
            })

            # Set the full path for the Excel file
            excel_file = os.path.join(output_path, file_name + '.xlsx')

            # Call the save function with the DataFrame and file path
            time.sleep(1)
            save_excel_file(output_df, excel_file, original_stdout)

            # Optionally, open the saved file
            os.startfile(output_path)

        # sys.stdout = original_stdout  # Reset stdout back to original

    except Exception as e:
        # Initialize logging only when an error is caught
        logging, error_file = setup_logging(downloads_folder)
        logging.error(f"An error occurred: {e}", exc_info=True)
        os.startfile(error_file)

def main():
    global log_file_entry, length_tol_entry, width_tol_entry, height_tol_entry, box_df, checkboxes, selected_boxes, tolerances

    # Create the main window
    root = tk.Tk()
    root.title("Log Parser")
    
    # Make the window stay on top of other windows
    root.attributes('-topmost', True)

    # Set a fixed window size
    root.geometry("500x500")  # You can adjust the size as needed

    # Create the label with underlined text and center it across all 3 columns
    label_font = font.Font(underline=True)  # Create a font object with underlined tex

    # Validation to ensure only numeric values (including decimals) are entered
    vcmd = (root.register(validate_numeric_input), '%d', '%P')

    # Create a frame for importing a log file
    frame_import = tk.Frame(root, padx=10, pady=10)
    frame_import.grid(row=0, column=0, columnspan=3, sticky="ew")

    import_button = tk.Button(frame_import, text="Import Log File", command=load_file)
    import_button.grid(row=0, column=0, padx=5)

    log_file_entry = tk.Entry(frame_import, text="No file selected.", width=55)
    log_file_entry.grid(row=0, column=1, padx=10, sticky="w")

    # Create a main frame to hold tolerances and boxes
    frame_main = tk.Frame(root, padx=10, pady=10)
    frame_main.grid(row=1, column=0, columnspan=3, sticky="ew")

    # Create sub-frame for tolerances (top side)
    frame_tolerances = tk.Frame(frame_main)
    frame_tolerances.grid(row=0, column=0, columnspan=3, sticky="ew")

    tk.Label(frame_tolerances, text="Enter Tolerances:", font=label_font).grid(row=0, column=0, sticky=tk.W)

    tolerances = load_values()

    # Length tolerance
    tk.Label(frame_tolerances, text="Length :\t\t±").grid(row=1, column=0, sticky=tk.W)
    length_tol_entry = tk.Entry(frame_tolerances, width=5, validate="key", validatecommand=vcmd)
    length_tol_entry.insert(0, tolerances["length"])
    length_tol_entry.grid(row=1, column=1, pady=5)
    
    # Width tolerance
    tk.Label(frame_tolerances, text="Width  :\t\t±").grid(row=2, column=0, sticky=tk.W)
    width_tol_entry = tk.Entry(frame_tolerances, width=5, validate="key", validatecommand=vcmd)
    width_tol_entry.insert(0, tolerances["width"])
    width_tol_entry.grid(row=2, column=1, pady=5)
    
    # Height tolerance
    tk.Label(frame_tolerances, text="Height :\t\t±").grid(row=3, column=0, sticky=tk.W)
    height_tol_entry = tk.Entry(frame_tolerances, width=5, validate="key", validatecommand=vcmd)
    height_tol_entry.insert(0, tolerances["height"])
    height_tol_entry.grid(row=3, column=1, pady=5)

    # Load the CSV with actual dimensions
    box_df = pd.read_csv(resource_path('data/Xactual.csv'))

    # Get unique box sizes from the "Box" column
    unique_boxes = box_df['Box'].unique()

    # Load previously saved selected boxes (if available)
    previous_selected_boxes = load_selected_boxes()

    # Create a dictionary to hold checkboxes
    checkboxes = {}

    # Create sub-frame for tolerances (top side)
    frame_boxes = tk.Frame(frame_main)
    frame_boxes.grid(row=2, column=0, pady=10, columnspan=3, sticky="ew")

    tk.Label(frame_boxes, text="Check Boxes Ran:", font=label_font).grid(row=0, column=0, sticky=tk.W)

    # Determine the number of rows needed for 3 columns
    num_boxes = len(unique_boxes)
    num_columns = 3
    num_rows = (num_boxes + num_columns - 1) // num_columns  # Calculate the number of rows

    # Add a checkbox for each unique box size
    for i, box in enumerate(unique_boxes):
        var = tk.StringVar(value='1' if previous_selected_boxes and box in previous_selected_boxes else '0')  # Set default based on saved data
        box_row = i % num_rows + 1  # Calculate row based on index
        box_column = i // num_rows # Calculate column based on index
        checkbox = ttk.Checkbutton(frame_boxes, text=box, variable=var)
        checkbox.grid(row=box_row, column=box_column, sticky='w', padx=5, pady=5)  # Place each checkbox in the correct row and column
        checkboxes[box] = var

    run_button = tk.Button(frame_boxes, text="Filter Boxes", command=lambda: [parse_log(), store_values(), root.quit()])
    run_button.grid(row=box_row+2, column=1, columnspan=3, sticky="ew")

    # Run the Tkinter event loop
    root.mainloop()

if __name__ == "__main__":
    # Run the main program
    main()